"""
두 엑셀 파일을 비교합니다.
"""

import os
import pandas as pd
from openpyxl import load_workbook


def add_seqno(excel_1_path: str, excel_2_path: str) -> pd.DataFrame:
    """두 엑셀 파일을 비교하여 위원명, BOOKID, 질의가 같을 경우 SEQNO을 추가합니다"""
    #! 위원명, BOOK_ID, SEQNO, 질의
    df1 = pd.read_excel(excel_1_path)
    df2 = pd.read_excel(excel_2_path, sheet_name=1)

    df1_subset = df1.iloc[:, [2, 3, 4, 5]]
    df2_subset = df2.iloc[:, [6, 3, 4, 7]]

    for _, row1 in df1_subset.iterrows():
        for idx2, row2 in df2_subset.iterrows():
            if (row1.iloc[1] == row2.iloc[1] and
                (pd.notna(row1.iloc[3]) and str(row1.iloc[3]).strip() ==
                 str(row2.iloc[3]).strip()) and pd.notna(row2.iloc[0])):
                df2.iloc[idx2, 4] = row1.iloc[2]

    return df2


def insert_filename_data(input_path, file_id):
    """3. BOOKID, SEQNO를 모두 병합 후, 순서대로 FILENAME을 삽입합니다"""
    wb = load_workbook(input_path)
    ws = wb.worksheets[1]
    wb.active = ws

    file_id_length = len(file_id)
    file_id_to_int = int(file_id)

    for ws_row_num in range(2, ws.max_row + 1):
        realfile_name = ws.cell(row=ws_row_num, column=11).value
        if realfile_name is None:
            continue
        _, extension = os.path.splitext(realfile_name)
        upper_extension = extension.upper()
        file_name =\
            f"{str(file_id_to_int).zfill(file_id_length)}{upper_extension}"

        ws.cell(row=ws_row_num, column=6, value=file_name)
        file_id_to_int += 1

    wb.save(input_path)
