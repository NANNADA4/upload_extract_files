"""
두 엑셀 파일을 비교합니다.
"""


from collections import OrderedDict
from openpyxl import Workbook


def compare_excel(excel_1=Workbook, excel_2=Workbook) -> Workbook:  # pylint: disable=R0914
    """두 엑셀 파일을 비교하여 값이 같을 경우 질의를 추가합니다"""
    cmp1_compare = [1, 2, 3, 6]  # 위원회, 피감기관, 위원명, 질의
    cmp2_compare = [1, 2, 4, 5]  # 위원회, 피감기관, 위원명, 질의
    total_dic = {}

    ws1 = excel_1.active
    ws2 = excel_2.active

    for ws1_row_num in range(2, ws1.max_row + 1):
        attach_cnt = 0
        attach_excel1_info_list = []
        attach_list = []
        cmp1_value = [ws1.cell(
            row=ws1_row_num, column=col).value for col in cmp1_compare]
        for ws2_row_num in range(2, ws2.max_row + 1):
            cmp2_value = [ws2.cell(
                row=ws2_row_num, column=col).value for col in cmp2_compare]

            if cmp1_value == cmp2_value:  # 질의들 리스트에 저장
                attach_list.append(ws2.cell(row=ws2_row_num, column=6).value)
                attach_cnt += 1

            if len(attach_list) == 0:
                continue

        # 위원회, 피감기관, 위원명, BOOKID, SEQNO, 질의, 별첨리스트
        if len(attach_list) > 0:
            attach_excel1_info_list.append([ws1.cell(row=ws1_row_num, column=1).value,
                                            ws1.cell(row=ws1_row_num, column=2).value, ws1.cell(
                                            row=ws1_row_num, column=3).value,
                                            ws1.cell(row=ws1_row_num, column=4).value, ws1.cell(
                                            row=ws1_row_num, column=5).value,
                                            ws1.cell(row=ws1_row_num, column=6).value, attach_list])

        # 행 번호, 리스트 dic. 엑셀 삽입시 역순으로
            total_dic.update(
                {ws1_row_num + 1: attach_excel1_info_list})

    # 엑셀 삽입
    sorted_keys = sorted(total_dic.keys(), reverse=True)
    final_dict = OrderedDict((key, total_dic[key]) for key in sorted_keys)

    for key, value in final_dict.items():
        ws1.insert_rows(key, len(value[0][6]))
        for cnt in range(len(value[0][6])):
            ws1.cell(row=key + cnt, column=1, value=value[0][0])
            ws1.cell(row=key + cnt, column=2, value=value[0][1])
            ws1.cell(row=key + cnt, column=3, value=value[0][2])
            ws1.cell(row=key + cnt, column=4, value=value[0][3])
            ws1.cell(row=key + cnt, column=5, value=value[0][4])
            ws1.cell(row=key + cnt, column=6, value=value[0][5])
            ws1.cell(row=key + cnt, column=7, value=value[0][6][cnt])

    return excel_1
