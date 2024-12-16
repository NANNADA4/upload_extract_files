"""
위원회, 피감기관, 위원명, 파일명을 비교하여 전부 일치하면 개인정보가 존재한다는 표시를 합니다
"""


import openpyxl


def combine_is_exist_personal_info(file_path):
    """위원회, 피감기관, 위원명, 파일명을 비교하여 전부 일치하면 개인정보가 존재한다는 표시를 합니다"""
    workbook = openpyxl.load_workbook(file_path)

    sheet1 = workbook.worksheets[0]  # 업로드 리스트
    sheet2 = workbook.worksheets[1]  # 개인정보 검출목록

    for row1 in range(2, sheet1.max_row + 1):
        val_a1 = sheet1.cell(row=row1, column=12).value  # L열
        space_removed_val_a1 = val_a1.replace(
            " ", "") if val_a1 else ""
        personal_info_found = False

        for row2 in range(2, sheet2.max_row + 1):
            val_a2 = sheet2.cell(row=row2, column=1).value
            space_removed_val_a2 = val_a2.replace(
                " ", "") if val_a2 else ""

            if space_removed_val_a1 in space_removed_val_a2:
                sheet1.cell(row=row1, column=17).value = sheet2.cell(
                    row=row2, column=2).value  # Q열
                personal_info_found = True
                break

        if not personal_info_found:
            sheet1.cell(row=row1, column=17).value = ""  # Q열

    workbook.save(file_path)
