"""
폴더를 순회하면서 파일을 찾고 파일명을 변경하고 폴더를 생성합니다.
"""


import os
import shutil
from datetime import datetime
import traceback
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


from module.__process__.process_log import create_log, get_log_path
from module.create_excel.create_excel import create_excel
from module.merge_excel.compare_excel import add_pdf_answer, insert_filename_data
from module.utils.combine_personal_info import combine_is_exist_personal_info


def take_exception(e):
    """exception 발생시 출력"""
    print(f"\n!!!!!{e}!!!!!\n")
    print("-"*10)
    traceback.print_exc()
    print("-"*10)
    print("")


def process_create(input_path, excel_path):
    """#1. BOOK_ID, SEQNO가 매겨진 엑셀파일을 생성합니다"""
    try:
        create_excel(input_path, excel_path)
        print("\n=> 엑셀 파일이 정상적으로 생성되었습니다\n")
    except Exception as e:  # pylint: disable=W0718
        take_exception(e)
        if os.path.exists(excel_path):
            os.remove(excel_path)


def get_time():
    """현재 시간, 분, 초를 반환합니다"""
    now = datetime.now()
    time_now = str(now.month).zfill(2) + str(now.day).zfill(2) + \
        str(now.hour).zfill(2) + str(now.minute).zfill(2)

    return time_now


def process_merge(base_excel_path, attach_excel_path):
    """#2. 1번에서 제작한 엑셀파일과 별도제출자료를 정리한 엑셀파일을 병합합니다"""
    try:
        df_merge = add_pdf_answer(base_excel_path, attach_excel_path)
        df_merge.to_excel(os.path.join(os.path.dirname(
            base_excel_path), f'업로드리스트_{get_time()}.xlsx'), index=False)
        print("\n=> SEQNO 병합이 완료되었습니다. \n")
    except Exception as e:  # pylint: disable=W0718
        take_exception(e)


def process_count(input_path, file_id):
    """#3. BOOKID, SEQNO를 모두 병합 후, 순서대로 FILENAME을 삽입합니다"""
    insert_filename_data(input_path, file_id)
    print("\n=> FILENAME이 모두 입력되었습니다.\n")


def process_rename(input_path, output_path, excel_path):
    """#4. 폴더를 순회하면서 엑셀 파일 내부 파일명을 변경하고 새로운 폴더로 이동합니다."""
    df = pd.read_excel(excel_path, engine='openpyxl', sheet_name=1)

    if not all(col in df.columns for col in ['실제 파일명', 'FILE_NAME', 'FILE_PATH']):
        print("엑셀 읽기 오류! : 엑셀 파일에 필요한 열이 없습니다.")
        return

    df['FILE_PATH'] = df['FILE_PATH'].astype(str)
    df['BOOKID'] = df['BOOKID'].astype(str)

    for index, row in df.iterrows():
        file_path = row['경로']
        actual_file_path = os.path.join(input_path, file_path)
        if not os.path.exists(actual_file_path):
            create_log(actual_file_path)
            continue
        file_path_row = f"/audit/attach/2023/{str(row['BOOKID']).zfill(9)}/" + \
            f"{str(row['FILE_NAME'])}"
        real_file_path = os.path.join(
            "inspection", "reqdoc", "2023", str(
                row['BOOKID']).zfill(9), str(row['FILE_NAME']))
        target_file_path = os.path.join(output_path, real_file_path)

        target_folder = os.path.dirname(target_file_path)
        if not os.path.exists(target_folder):
            os.makedirs(target_folder)

        if os.path.exists(actual_file_path):
            shutil.copy(actual_file_path, target_file_path)

        df.at[index, 'FILE_PATH'] = file_path_row
        df.at[index, 'FILE_NAME'] = row['FILE_NAME']

    print("\n~~~엑셀 파일 수정중입니다~~~")

    try:
        df.to_excel(os.path.join(os.path.dirname(excel_path),
                                 f"최종업로드리스트_{get_time()}.xlsx"), index=False, engine='openpyxl')
    except PermissionError:
        print("엑셀 수정 실패! : 엑셀 파일이 열려있는 경우, 닫고 다시 실행하세요")
        return


def log_to_excel(output_path):
    """로그파일을 토대로 엑셀을 생성합니다"""
    if not os.path.exists(get_log_path()):
        print("=> 로그 파일이 존재하지 않습니다\n")
        return

    wb = Workbook()
    ws = wb.active
    headers = ['연번', '파일명', '확장자', '경로']
    header_color = PatternFill(start_color='4f81bd',
                               end_color='4f81bd', fill_type='solid')
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=1, column=col_idx).fill = header_color

    with open(get_log_path(), 'r', encoding='UTF-8') as file:
        lines = file.readlines()
    max_low = ws.max_row + 1
    for idx, line in enumerate(lines):
        filename, extension = os.path.splitext(os.path.basename(line))
        extension_wo_dot = extension.lstrip('.')
        ws.cell(row=max_low, column=1, value=idx + 1)
        ws.cell(row=max_low, column=2, value=filename)
        ws.cell(row=max_low, column=3, value=extension_wo_dot)
        ws.cell(row=max_low, column=4, value=line)
        max_low += 1

    wb.save(output_path)


def process_folder(input_num) -> bool:
    """전달받은 input_num을 토대로 알맞은 함수로 반환합니다."""
    if input_num in ['1', '4']:
        input_path = input("입력 폴더의 경로를 입력하세요\n=> ")
        input_path = os.path.join('\\\\?\\', input_path)

        if not os.path.isdir(input_path):
            print("\n====입력 폴더의 경로를 다시 한 번 확인하세요====\n")
            return False

    match input_num:
        case '1':
            while True:
                create_excel_path = input("엑셀 파일을 저장할 경로를 입력하세요\n=> ")
                if not create_excel_path.lower().endswith('.xlsx'):
                    print("====엑셀 파일 확장자를 입력했는지 확인해주세요====")
                process_create(input_path, create_excel_path)
                break
        case '2':
            while True:
                base_excel_path = input("1번에서 실행한 결과의 엑셀파일 경로를 입력하세요\n=> ")
                attach_excel_path = input("별도제출자료를 정리한 엑셀파일 경로를 입력하세요\n=> ")
                if os.path.exists(base_excel_path) and os.path.exists(attach_excel_path):
                    process_merge(base_excel_path, attach_excel_path)
                    break
                print("\n=> 엑셀 파일 경로를 한번 더 확인해주세요")
        case '3':
            while True:
                count_excel_path = input("엑셀파일 경로를 입력해주세요\n=> ")
                file_id = input("FILE_NAME에 들어갈 시작번호를 입력하세요 (seqNO순)\n=> ")
                try:
                    int(file_id)
                except ValueError:
                    print("\n====숫자만 입력해주세요====\n")
                if os.path.exists(count_excel_path):
                    process_count(count_excel_path, file_id)
                    break
        case '4':
            rename_output_path = input(
                "경로 및 이름을 변경한 파일을 복사할 폴더 경로를 입력하세요\n=> ")
            rename_excel_path = input(
                "엑셀 파일의 경로를 입력하세요 (엑셀이 종료되었는지 확인하세요)\n=> ")
            process_rename(input_path, rename_output_path, rename_excel_path)
        case '5':
            output_path = input("저장할 엑셀파일 경로를 입력하세요\n=> ")
            log_to_excel(output_path)
            print("엑셀 파일 변환이 완료되었습니다.")
        case '6':
            file_path = input("엑셀 파일 경로를 입력해주세요 (확장자 포함)\n=> ")
            combine_is_exist_personal_info(file_path)

    return True
